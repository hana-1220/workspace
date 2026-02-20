#!/usr/bin/env python3
"""
SquadBeyond ヒートマップ分析 Excel レポート生成スクリプト

使い方:
  1. config と blocks を分析結果に合わせて書き換える
  2. python heatmap_analysis.py を実行
  3. output/ フォルダに Excel が生成される
"""

import math
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 設定（分析ごとに書き換える）
# ============================================================
config = {
    "page_name": "サンプルLP",
    "version": "v1",
    "period": "（SquadBeyondスクショ取得日: 2026/02/20）",
    "cv_users": 0,
    "click_users": 0,
    "exit_users": 0,
    "findings": [
        "【発見1】（3列比較のインサイトをここに記述）",
        "【発見2】（3列比較のインサイトをここに記述）",
        "【発見3】（3列比較のインサイトをここに記述）",
    ],
}

# ============================================================
# ブロックデータ（ヒートマップスクショから読み取り）
# ============================================================
blocks = [
    # テンプレート:
    # {
    #     "name": "ブロック名",
    #     "position": "0-5%",
    #     "cv_dwell": 3, "click_dwell": 3, "exit_dwell": 3,
    #     "cum_exit": 5,
    #     "click_resp": "△", "cv_resp": "×",
    #     "memo": "分析メモ",
    #     "action": "★ 施策案",
    # },
]

# ============================================================
# スタイル定数
# ============================================================
DWELL_FILLS = {
    5: PatternFill("solid", fgColor="FF3B30"),
    4: PatternFill("solid", fgColor="FF9500"),
    3: PatternFill("solid", fgColor="FFCC00"),
    2: PatternFill("solid", fgColor="A8D5BA"),
    1: PatternFill("solid", fgColor="5AC8FA"),
}
DWELL_FONTS = {
    5: Font(color="FFFFFF", bold=True),
    4: Font(color="FFFFFF", bold=True),
    3: Font(color="000000", bold=True),
    2: Font(color="000000", bold=True),
    1: Font(color="000000", bold=True),
}
DWELL_LABELS = {5: "🔴", 4: "🟠", 3: "🟡", 2: "🟢", 1: "🔵"}

EXIT_FILLS = {
    "4+": PatternFill("solid", fgColor="C0392B"),
    "3": PatternFill("solid", fgColor="E74C3C"),
    "2": PatternFill("solid", fgColor="F39C12"),
    "0-1": PatternFill("solid", fgColor="27AE60"),
}
EXIT_FONTS = {
    "4+": Font(color="FFFFFF", bold=True),
    "3": Font(color="FFFFFF", bold=True),
    "2": Font(color="000000", bold=True),
    "0-1": Font(color="FFFFFF", bold=True),
}

PRIORITY_FILLS = {
    "🔴最優先": PatternFill("solid", fgColor="FFCCCC"),
    "🟡要改善": PatternFill("solid", fgColor="FFF3CD"),
    "🟢維持": PatternFill("solid", fgColor="D4EDDA"),
}

INTEREST_FILLS = {
    "2.5+": PatternFill("solid", fgColor="1B5E20"),
    "2.0+": PatternFill("solid", fgColor="4CAF50"),
    "1.5+": PatternFill("solid", fgColor="FFC107"),
    "<1.5": PatternFill("solid", fgColor="FF5722"),
}
INTEREST_FONTS = {
    "2.5+": Font(color="FFFFFF", bold=True),
    "2.0+": Font(color="FFFFFF", bold=True),
    "1.5+": Font(color="000000", bold=True),
    "<1.5": Font(color="FFFFFF", bold=True),
}

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")


# ============================================================
# 計算ロジック
# ============================================================
RESP_SCORE = {"◎": 3, "○": 2, "△": 1, "×": 0}


def calc_weights(cfg):
    """CV / CLICK / 離脱 の人数比を返す。不明時はデフォルト比率。"""
    cv = cfg.get("cv_users", 0)
    click = cfg.get("click_users", 0)
    exit_ = cfg.get("exit_users", 0)
    total = cv + click + exit_
    if total == 0:
        return 0.13, 0.30, 0.57
    return cv / total, click / total, exit_ / total


def calc_dwell_overall(block, w_cv, w_click, w_exit):
    """滞在（総合）= 人数加重平均"""
    val = (
        block["cv_dwell"] * w_cv
        + block["click_dwell"] * w_click
        + block["exit_dwell"] * w_exit
    )
    return round(val, 2)


def calc_dwell_rationale(block, w_cv, w_click, w_exit):
    """滞在（総合）の根拠文字列"""
    return (
        f"CV{block['cv_dwell']}×{w_cv:.0%}"
        f" + CLICK{block['click_dwell']}×{w_click:.0%}"
        f" + 離脱{block['exit_dwell']}×{w_exit:.0%}"
    )


def calc_section_exit(blocks_list):
    """区間離脱率を算出して各blockに追加"""
    prev = 0
    for b in blocks_list:
        b["section_exit"] = round(b["cum_exit"] - prev, 1)
        prev = b["cum_exit"]


def exit_risk_key(section_exit):
    """区間離脱率 → リスクキー"""
    if section_exit >= 4:
        return "4+"
    elif section_exit >= 3:
        return "3"
    elif section_exit >= 2:
        return "2"
    else:
        return "0-1"


def exit_risk_label(section_exit):
    """区間離脱率 → リスクラベル"""
    mapping = {"4+": "🔴高", "3": "🟠やや高", "2": "🟡中", "0-1": "🟢低"}
    return mapping[exit_risk_key(section_exit)]


def calc_interest(block):
    """関心度スコア"""
    val = (
        block["cv_dwell"] * 0.2
        + block["click_dwell"] * 0.2
        + RESP_SCORE[block["click_resp"]] * 0.3
        + RESP_SCORE[block["cv_resp"]] * 0.3
    )
    return round(val, 2)


def interest_fill_key(score):
    """関心度 → 色キー"""
    if score >= 2.5:
        return "2.5+"
    elif score >= 2.0:
        return "2.0+"
    elif score >= 1.5:
        return "1.5+"
    else:
        return "<1.5"


def calc_priority(block, interest):
    """改善優先度を判定"""
    se = block["section_exit"]
    if se >= 4:
        return "🔴最優先"
    if se >= 2 and interest < 1.5:
        return "🔴最優先"
    if se >= 2:
        return "🟡要改善"
    if interest >= 2.0:
        return "🟢維持"
    return "🟡要改善"


# ============================================================
# 自動行高さ（日本語幅考慮）
# ============================================================
def _text_width_estimate(text, font_size=10):
    """日本語混在テキストの概算幅（ポイント単位）"""
    width = 0
    for ch in str(text):
        if ord(ch) > 0x7F:
            width += font_size * 1.8
        else:
            width += font_size * 0.7
    return width


def auto_row_heights(ws):
    """全行の高さをセル内容に合わせて自動調整"""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        max_lines = 1
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            col_width_chars = ws.column_dimensions[
                get_column_letter(cell.column)
            ].width or 12
            col_width_pt = col_width_chars * 7
            text_w = _text_width_estimate(text)
            lines_by_width = math.ceil(text_w / col_width_pt) if col_width_pt > 0 else 1
            lines_by_newline = text.count("\n") + 1
            max_lines = max(max_lines, lines_by_width, lines_by_newline)
        ws.row_dimensions[row[0].row].height = max(18, max_lines * 18)


# ============================================================
# Sheet1: ブロック別分析
# ============================================================
SHEET1_HEADERS = [
    ("No", 5),
    ("ブロック名", 22),
    ("位置", 9),
    ("CV滞在", 9),
    ("CLICK滞在", 10),
    ("離脱滞在", 9),
    ("滞在総合", 9),
    ("根拠", 30),
    ("累計離脱%", 10),
    ("区間離脱%", 10),
    ("離脱リスク", 10),
    ("Click反応", 10),
    ("CV反応", 9),
    ("関心度", 9),
    ("改善優先度", 12),
    ("分析メモ", 45),
    ("施策案", 35),
]


def build_sheet1(wb, blocks_list, w_cv, w_click, w_exit):
    ws = wb.active
    ws.title = "ブロック別分析"
    ws.sheet_properties.tabColor = "2C3E50"

    # --- ページ情報ヘッダー ---
    ws.merge_cells("A1:Q1")
    info_cell = ws["A1"]
    info_cell.value = (
        f"{config['page_name']}　{config['version']}　"
        f"期間: {config['period']}　"
        f"CV: {config['cv_users']:,}人 / CLICK: {config['click_users']:,}人 / "
        f"離脱: {config['exit_users']:,}人"
    )
    info_cell.font = Font(bold=True, size=12)
    info_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30

    # --- 列幅 & ヘッダー ---
    for ci, (name, width) in enumerate(SHEET1_HEADERS, 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=2, column=ci, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER

    # --- データ行 ---
    for ri, b in enumerate(blocks_list, 3):
        idx = ri - 2
        overall = calc_dwell_overall(b, w_cv, w_click, w_exit)
        rationale = calc_dwell_rationale(b, w_cv, w_click, w_exit)
        interest = calc_interest(b)
        priority = calc_priority(b, interest)

        row_data = [
            idx,
            b["name"],
            b["position"],
            b["cv_dwell"],
            b["click_dwell"],
            b["exit_dwell"],
            overall,
            rationale,
            b["cum_exit"],
            b["section_exit"],
            exit_risk_label(b["section_exit"]),
            b["click_resp"],
            b["cv_resp"],
            interest,
            priority,
            b.get("memo", ""),
            b.get("action", ""),
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

        # 滞在セル色
        for ci, key in [(4, "cv_dwell"), (5, "click_dwell"), (6, "exit_dwell")]:
            v = b[key]
            cell = ws.cell(row=ri, column=ci)
            cell.value = f"{DWELL_LABELS[v]} {v}"
            cell.fill = DWELL_FILLS[v]
            cell.font = DWELL_FONTS[v]
            cell.alignment = CENTER_WRAP

        # 滞在総合セル色（最も近い整数で色付け）
        overall_rounded = max(1, min(5, round(overall)))
        cell_overall = ws.cell(row=ri, column=7)
        cell_overall.fill = DWELL_FILLS[overall_rounded]
        cell_overall.font = DWELL_FONTS[overall_rounded]
        cell_overall.alignment = CENTER_WRAP

        # 離脱リスク色
        ek = exit_risk_key(b["section_exit"])
        cell_risk = ws.cell(row=ri, column=11)
        cell_risk.fill = EXIT_FILLS[ek]
        cell_risk.font = EXIT_FONTS[ek]
        cell_risk.alignment = CENTER_WRAP

        # 関心度色
        ik = interest_fill_key(interest)
        cell_int = ws.cell(row=ri, column=14)
        cell_int.fill = INTEREST_FILLS[ik]
        cell_int.font = INTEREST_FONTS[ik]
        cell_int.alignment = CENTER_WRAP

        # 優先度行背景
        if priority in PRIORITY_FILLS:
            pf = PRIORITY_FILLS[priority]
            cell_pri = ws.cell(row=ri, column=15)
            cell_pri.fill = pf

        # 反応セル中央揃え
        for ci in (12, 13):
            ws.cell(row=ri, column=ci).alignment = CENTER_WRAP

    # フリーズペイン
    ws.freeze_panes = "C3"

    auto_row_heights(ws)
    return ws


# ============================================================
# Sheet2: サマリー
# ============================================================
def build_sheet2(wb, blocks_list, w_cv, w_click, w_exit):
    ws = wb.create_sheet("サマリー")
    ws.sheet_properties.tabColor = "2980B9"

    title_font = Font(bold=True, size=14, color="2C3E50")
    section_font = Font(bold=True, size=12, color="2C3E50")
    section_fill = PatternFill("solid", fgColor="ECF0F1")

    # --- タイトル ---
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"ヒートマップ分析サマリー — {config['page_name']} {config['version']}"
    c.font = title_font
    ws.row_dimensions[1].height = 35

    # --- 基本情報 ---
    row = 3
    ws.cell(row=row, column=1, value="分析期間").font = Font(bold=True)
    ws.cell(row=row, column=2, value=config["period"])
    row += 1
    ws.cell(row=row, column=1, value="ユーザー数").font = Font(bold=True)
    ws.cell(
        row=row,
        column=2,
        value=(
            f"CV: {config['cv_users']:,} / "
            f"CLICK: {config['click_users']:,} / "
            f"離脱: {config['exit_users']:,}"
        ),
    )

    # --- 3列滞在パターン比較表 ---
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sec = ws.cell(row=row, column=1, value="■ 3列の滞在パターン比較")
    sec.font = section_font
    sec.fill = section_fill

    row += 1
    compare_headers = ["ブロック名", "CV滞在", "CLICK滞在", "離脱滞在", "差分メモ"]
    for ci, h in enumerate(compare_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER_WRAP
        c.border = THIN_BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 30

    for b in blocks_list:
        row += 1
        vals = [b["name"], b["cv_dwell"], b["click_dwell"], b["exit_dwell"]]
        diff = b["cv_dwell"] - b["exit_dwell"]
        if diff >= 2:
            diff_memo = f"CV vs 離脱で差{diff} → CVユーザーが特に注視"
        elif diff <= -2:
            diff_memo = f"離脱 vs CVで差{abs(diff)} → 離脱層が滞留（迷い？）"
        else:
            diff_memo = "3列で大きな差なし"
        vals.append(diff_memo)
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER
        # 滞在セル色
        for ci, dval in [(2, b["cv_dwell"]), (3, b["click_dwell"]), (4, b["exit_dwell"])]:
            c = ws.cell(row=row, column=ci)
            c.fill = DWELL_FILLS[dval]
            c.font = DWELL_FONTS[dval]
            c.alignment = CENTER_WRAP

    # --- TOP改善アクション ---
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sec = ws.cell(row=row, column=1, value="■ TOP改善アクション")
    sec.font = section_font
    sec.fill = section_fill

    row += 1
    act_headers = ["優先度", "ブロック", "区間離脱%", "関心度", "施策案"]
    for ci, h in enumerate(act_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER_WRAP
        c.border = THIN_BORDER

    sorted_blocks = sorted(
        blocks_list,
        key=lambda x: (
            0 if calc_priority(x, calc_interest(x)) == "🔴最優先" else
            1 if calc_priority(x, calc_interest(x)) == "🟡要改善" else 2,
            -x["section_exit"],
        ),
    )
    for b in sorted_blocks:
        interest = calc_interest(b)
        priority = calc_priority(b, interest)
        row += 1
        vals = [priority, b["name"], b["section_exit"], interest, b.get("action", "")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER
        if priority in PRIORITY_FILLS:
            ws.cell(row=row, column=1).fill = PRIORITY_FILLS[priority]

    # --- 最大の発見 ---
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sec = ws.cell(row=row, column=1, value="■ 最大の発見（3列比較インサイト）")
    sec.font = section_font
    sec.fill = section_fill

    findings = config.get("findings", [])
    for finding in findings:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=finding)
        c.alignment = WRAP_ALIGN
        c.border = THIN_BORDER

    auto_row_heights(ws)
    return ws


# ============================================================
# Sheet3: 離脱カーブ
# ============================================================
def build_sheet3(wb, blocks_list):
    ws = wb.create_sheet("離脱カーブ")
    ws.sheet_properties.tabColor = "E74C3C"

    headers = ["No", "ブロック名", "位置", "累計離脱%", "残留率%", "区間離脱%"]
    col_widths = [5, 22, 10, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER_WRAP
        c.border = THIN_BORDER

    for ri, b in enumerate(blocks_list, 2):
        idx = ri - 1
        retention = round(100 - b["cum_exit"], 1)
        vals = [idx, b["name"], b["position"], b["cum_exit"], retention, b["section_exit"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = WRAP_ALIGN
            c.border = THIN_BORDER

    # --- 離脱カーブチャート ---
    if len(blocks_list) >= 2:
        chart = LineChart()
        chart.title = "離脱カーブ（累計離脱% / 残留率%）"
        chart.y_axis.title = "%"
        chart.x_axis.title = "ブロック"
        chart.style = 10
        chart.width = 28
        chart.height = 14

        max_row = len(blocks_list) + 1
        cats = Reference(ws, min_col=3, min_row=2, max_row=max_row)

        cum_data = Reference(ws, min_col=4, min_row=1, max_row=max_row)
        chart.add_data(cum_data, titles_from_data=True)

        ret_data = Reference(ws, min_col=5, min_row=1, max_row=max_row)
        chart.add_data(ret_data, titles_from_data=True)

        chart.set_categories(cats)

        s0 = chart.series[0]
        s0.graphicalProperties.line.solidFill = "E74C3C"
        s1 = chart.series[1]
        s1.graphicalProperties.line.solidFill = "27AE60"

        ws.add_chart(chart, f"A{max_row + 2}")

    auto_row_heights(ws)
    return ws


# ============================================================
# Sheet4: スコア算出ロジック
# ============================================================
def build_sheet4(wb):
    ws = wb.create_sheet("スコア算出ロジック")
    ws.sheet_properties.tabColor = "8E44AD"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    title_font = Font(bold=True, size=12, color="2C3E50")
    section_fill = PatternFill("solid", fgColor="ECF0F1")

    rows = [
        ("■ 滞在時間スケール", "", True),
        ("5 🔴", "長い（赤）— ユーザーが最も長く注視", False),
        ("4 🟠", "やや長い（オレンジ）", False),
        ("3 🟡", "普通（黄）", False),
        ("2 🟢", "やや短い（黄緑）", False),
        ("1 🔵", "短い（緑〜青）— ほぼ素通り", False),
        ("", "", False),
        ("■ 滞在（総合）", "", True),
        (
            "計算式",
            "CV滞在×(CV人数/全体) + CLICK滞在×(CLICK人数/全体) + 離脱滞在×(離脱人数/全体)",
            False,
        ),
        (
            "デフォルト比率",
            f"CV={config['cv_users']:,} / CLICK={config['click_users']:,} / "
            f"離脱={config['exit_users']:,}",
            False,
        ),
        ("", "", False),
        ("■ 関心度", "", True),
        (
            "計算式",
            "CV滞在×0.2 + CLICK滞在×0.2 + クリック反応×0.3 + CV反応×0.3",
            False,
        ),
        ("反応スコア", "◎=3, ○=2, △=1, ×=0", False),
        ("", "", False),
        ("■ 改善優先度", "", True),
        ("🔴最優先", "区間離脱4%以上 / 離脱高×関心低 / CTA前の致命的離脱", False),
        ("🟡要改善", "区間離脱2-3% / 列間でばらつきあり", False),
        ("🟢維持", "離脱低×関心高", False),
        ("", "", False),
        ("■ 離脱リスク", "", True),
        ("🔴高", "区間離脱 4%以上", False),
        ("🟠やや高", "区間離脱 3%台", False),
        ("🟡中", "区間離脱 2%台", False),
        ("🟢低", "区間離脱 0-1%", False),
        ("", "", False),
        ("■ 関心度レベル", "", True),
        ("2.5以上", "高関心（濃緑）", False),
        ("2.0以上", "中関心（緑）", False),
        ("1.5以上", "低関心（黄）", False),
        ("1.5未満", "無関心（赤橙）", False),
    ]

    for ri, (a, b, is_section) in enumerate(rows, 1):
        ca = ws.cell(row=ri, column=1, value=a)
        cb = ws.cell(row=ri, column=2, value=b)
        ca.alignment = WRAP_ALIGN
        cb.alignment = WRAP_ALIGN
        ca.border = THIN_BORDER
        cb.border = THIN_BORDER
        if is_section:
            ca.font = title_font
            ca.fill = section_fill
            cb.fill = section_fill

    auto_row_heights(ws)
    return ws


# ============================================================
# メイン
# ============================================================
def main():
    if not blocks:
        print("⚠️  blocks が空です。分析データを入力してから実行してください。")
        return

    # 区間離脱を算出
    calc_section_exit(blocks)

    # 重み
    w_cv, w_click, w_exit = calc_weights(config)

    # ワークブック作成
    wb = Workbook()

    build_sheet1(wb, blocks, w_cv, w_click, w_exit)
    build_sheet2(wb, blocks, w_cv, w_click, w_exit)
    build_sheet3(wb, blocks)
    build_sheet4(wb)

    # 出力
    os.makedirs("output", exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"output/heatmap_{config['page_name']}_{config['version']}_{ts}.xlsx"
    wb.save(filename)
    print(f"✅ Excel出力完了: {filename}")


if __name__ == "__main__":
    main()
